import requests
from bs4 import BeautifulSoup
import lxml
import fake_useragent
import openpyxl


user = fake_useragent.UserAgent().random

headers = {"user-agent": user}
session = requests.Session()
book = openpyxl.Workbook()
sheet = book.active
for j in range(1, 14):
    url = f'https://allo.ua/ua/roboty-pylesosy/p-{j}/'
    response = session.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "lxml")
    paranc = soup.find('div', class_='products-layout__container products-layout--grid')
    product = paranc.find_all('div', class_='product-card')

    for i in range(len(product)):
        title_product = product[i].find('a', class_='product-card__title')
        try:
            price_product = product[i].find('span', class_='sum')
            result = [title_product.text,  price_product.text]
            sheet.cell(row=(i+1)+(j-1)*10, column=1).value = result[0]
            sheet.cell(row=(i+1)+(j-1)*10, column=2).value = result[1]
            # f = open('product.txt', 'a', encoding='utf-8')
            # f.write(title_product.text + ' ' + price_product.text + "\n")
            # f.close()
        except AttributeError:
            print('Нет в наличии')
    print(f'Page {j}')
book.save('product.xlsx')
book.close()